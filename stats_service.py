from __future__ import annotations

import os
import sqlite3
import threading
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable, Optional

from loguru import logger
from openpyxl import Workbook
from tzlocal import get_localzone

from hide_private_data import mask_sensitive_data


_INIT_LOCK = threading.Lock()
_WRITE_LOCK = threading.Lock()
_INITIALIZED: set[Path] = set()


def _utc_date_from_ts(ts: int) -> str:
    return datetime.fromtimestamp(int(ts), tz=timezone.utc).date().isoformat()


def _coerce_db_path(db_path: str | Path) -> Path:
    p = Path(db_path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    # Touch so we fail early on permissions.
    p.touch(exist_ok=True)
    return p


def resolve_default_db_path() -> Path:
    """Keep consistent with AvitoParse._resolve_db_path()."""
    default = Path(__file__).resolve().parent / "database.db"
    try:
        default.touch(exist_ok=True)
        if not os.access(default, os.W_OK | os.R_OK):
            raise PermissionError("no access to database.db")
        return default
    except PermissionError:
        import tempfile

        fallback = Path(tempfile.gettempdir()) / "avito_parser_database.db"
        fallback.touch(exist_ok=True)
        return fallback


def _ensure_schema(db_path: Path) -> None:
    with _INIT_LOCK:
        if db_path in _INITIALIZED:
            return
        conn = sqlite3.connect(str(db_path), timeout=30)
        try:
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS daily_stats (
                    date TEXT NOT NULL,
                    chat_id TEXT NOT NULL,
                    requests_total INTEGER NOT NULL DEFAULT 0,
                    ok_total INTEGER NOT NULL DEFAULT 0,
                    blocked_total INTEGER NOT NULL DEFAULT 0,
                    rate_limited_total INTEGER NOT NULL DEFAULT 0,
                    proxy_error_total INTEGER NOT NULL DEFAULT 0,
                    other_error_total INTEGER NOT NULL DEFAULT 0,
                    direct_total INTEGER NOT NULL DEFAULT 0,
                    user_proxy_total INTEGER NOT NULL DEFAULT 0,
                    free_proxy_total INTEGER NOT NULL DEFAULT 0,
                    items_total INTEGER NOT NULL DEFAULT 0,
                    last_ts INTEGER NOT NULL DEFAULT 0,
                    PRIMARY KEY(date, chat_id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS item_hits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ts INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    chat_id TEXT NOT NULL,
                    filter_id INTEGER,
                    filter_title TEXT,
                    item_id TEXT,
                    item_url TEXT,
                    title TEXT,
                    price INTEGER,
                    region TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_item_hits_date_chat ON item_hits(date, chat_id)")
            conn.commit()
        finally:
            conn.close()
        _INITIALIZED.add(db_path)


@dataclass(frozen=True)
class DailyStatsRow:
    date: str
    chat_id: str
    requests_total: int
    ok_total: int
    blocked_total: int
    rate_limited_total: int
    proxy_error_total: int
    other_error_total: int
    direct_total: int
    user_proxy_total: int
    free_proxy_total: int
    items_total: int
    last_ts: int


class StatsDB:
    """
    Lightweight request + item statistics.

    Goals:
    - be safe (never crash the parser on stats issues)
    - be fast (upsert counters)
    - be thread-friendly (single process with a few worker threads)
    """

    def __init__(self, db_path: str | Path | None = None, *, enabled: bool = True):
        disable = (os.getenv("AVITO_DISABLE_STATS") or "").strip().lower() in {"1", "true", "yes", "on"}
        self.enabled = bool(enabled) and (not disable)
        if not self.enabled:
            self.db_path = None
            return
        self.db_path = _coerce_db_path(db_path or resolve_default_db_path())
        _ensure_schema(self.db_path)

    def _connect(self) -> sqlite3.Connection:
        assert self.db_path is not None
        conn = sqlite3.connect(str(self.db_path), timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def record_request(
        self,
        *,
        chat_id: str,
        filter_id: int | None,
        filter_title: str | None,
        url: str,
        proxy_kind: str,  # direct|user|free
        status_code: int | None,
        outcome: str,  # ok|blocked|rate_limited|proxy_error|other_error
        duration_ms: int | None = None,
        scheduler_mode: bool = False,
        error: str | None = None,
        ts: int | None = None,
    ) -> None:
        if not self.enabled or not self.db_path:
            return
        try:
            ts_val = int(ts if ts is not None else time.time())
            date = _utc_date_from_ts(ts_val)
            chat = str(chat_id or "global")
            kind = (proxy_kind or "direct").strip().lower()
            if kind not in {"direct", "user", "free"}:
                kind = "direct"
            out = (outcome or "other_error").strip().lower()
            if out not in {"ok", "blocked", "rate_limited", "proxy_error", "other_error"}:
                out = "other_error"

            # Counter increments.
            ok_inc = 1 if out == "ok" else 0
            blocked_inc = 1 if out == "blocked" else 0
            rl_inc = 1 if out == "rate_limited" else 0
            proxy_err_inc = 1 if out == "proxy_error" else 0
            other_err_inc = 1 if out == "other_error" else 0
            direct_inc = 1 if kind == "direct" else 0
            user_inc = 1 if kind == "user" else 0
            free_inc = 1 if kind == "free" else 0

            # Never store secrets in error strings.
            safe_error = mask_sensitive_data(str(error))[:500] if error else None

            # Single upsert to keep it fast.
            with _WRITE_LOCK:
                conn = self._connect()
                try:
                    conn.execute(
                        """
                        INSERT INTO daily_stats(
                            date, chat_id,
                            requests_total, ok_total, blocked_total, rate_limited_total,
                            proxy_error_total, other_error_total,
                            direct_total, user_proxy_total, free_proxy_total,
                            items_total, last_ts
                        )
                        VALUES (?, ?,
                            1, ?, ?, ?,
                            ?, ?,
                            ?, ?, ?,
                            0, ?
                        )
                        ON CONFLICT(date, chat_id) DO UPDATE SET
                            requests_total = requests_total + 1,
                            ok_total = ok_total + excluded.ok_total,
                            blocked_total = blocked_total + excluded.blocked_total,
                            rate_limited_total = rate_limited_total + excluded.rate_limited_total,
                            proxy_error_total = proxy_error_total + excluded.proxy_error_total,
                            other_error_total = other_error_total + excluded.other_error_total,
                            direct_total = direct_total + excluded.direct_total,
                            user_proxy_total = user_proxy_total + excluded.user_proxy_total,
                            free_proxy_total = free_proxy_total + excluded.free_proxy_total,
                            last_ts = CASE WHEN excluded.last_ts > last_ts THEN excluded.last_ts ELSE last_ts END
                        """,
                        (
                            date,
                            chat,
                            ok_inc,
                            blocked_inc,
                            rl_inc,
                            proxy_err_inc,
                            other_err_inc,
                            direct_inc,
                            user_inc,
                            free_inc,
                            ts_val,
                        ),
                    )
                    # Optional: keep detailed request logs out of the DB (fast + small).
                    _ = (url, filter_id, filter_title, status_code, duration_ms, scheduler_mode, safe_error)
                    conn.commit()
                finally:
                    conn.close()
        except Exception as err:
            logger.debug("stats record_request failed: {}", str(err)[:160])

    def record_item(
        self,
        *,
        chat_id: str,
        filter_id: int | None,
        filter_title: str | None,
        item_id: str | None,
        item_url: str | None,
        title: str | None,
        price: int | None,
        region: str | None,
        ts: int | None = None,
    ) -> None:
        if not self.enabled or not self.db_path:
            return
        try:
            ts_val = int(ts if ts is not None else time.time())
            date = _utc_date_from_ts(ts_val)
            chat = str(chat_id or "global")
            with _WRITE_LOCK:
                conn = self._connect()
                try:
                    conn.execute(
                        """
                        INSERT INTO item_hits(
                            ts, date, chat_id, filter_id, filter_title,
                            item_id, item_url, title, price, region
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            ts_val,
                            date,
                            chat,
                            filter_id,
                            (filter_title or "")[:200],
                            (str(item_id) if item_id is not None else "")[:100],
                            (item_url or "")[:500],
                            (title or "")[:500],
                            int(price or 0),
                            (region or "")[:200],
                        ),
                    )
                    conn.execute(
                        """
                        INSERT INTO daily_stats(date, chat_id, items_total, last_ts)
                        VALUES (?, ?, 1, ?)
                        ON CONFLICT(date, chat_id) DO UPDATE SET
                            items_total = items_total + 1,
                            last_ts = CASE WHEN excluded.last_ts > last_ts THEN excluded.last_ts ELSE last_ts END
                        """,
                        (date, chat, ts_val),
                    )
                    conn.commit()
                finally:
                    conn.close()
        except Exception as err:
            logger.debug("stats record_item failed: {}", str(err)[:160])

    def get_daily_rows(self, chat_id: str, *, since_date: str | None = None) -> list[DailyStatsRow]:
        if not self.enabled or not self.db_path:
            return []
        chat = str(chat_id or "global")
        try:
            conn = self._connect()
            try:
                if since_date:
                    rows = conn.execute(
                        "SELECT * FROM daily_stats WHERE chat_id=? AND date>=? ORDER BY date ASC",
                        (chat, since_date),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT * FROM daily_stats WHERE chat_id=? ORDER BY date ASC",
                        (chat,),
                    ).fetchall()
            finally:
                conn.close()
            return [
                DailyStatsRow(
                    date=r["date"],
                    chat_id=r["chat_id"],
                    requests_total=int(r["requests_total"] or 0),
                    ok_total=int(r["ok_total"] or 0),
                    blocked_total=int(r["blocked_total"] or 0),
                    rate_limited_total=int(r["rate_limited_total"] or 0),
                    proxy_error_total=int(r["proxy_error_total"] or 0),
                    other_error_total=int(r["other_error_total"] or 0),
                    direct_total=int(r["direct_total"] or 0),
                    user_proxy_total=int(r["user_proxy_total"] or 0),
                    free_proxy_total=int(r["free_proxy_total"] or 0),
                    items_total=int(r["items_total"] or 0),
                    last_ts=int(r["last_ts"] or 0),
                )
                for r in rows
            ]
        except Exception as err:
            logger.debug("stats get_daily_rows failed: {}", str(err)[:160])
            return []

    def get_item_rows(self, chat_id: str, *, since_date: str | None = None, limit: int = 5000) -> list[dict]:
        if not self.enabled or not self.db_path:
            return []
        chat = str(chat_id or "global")
        try:
            conn = self._connect()
            try:
                if since_date:
                    rows = conn.execute(
                        """
                        SELECT ts, filter_id, filter_title, item_id, item_url, title, price, region
                        FROM item_hits
                        WHERE chat_id=? AND date>=?
                        ORDER BY ts DESC
                        LIMIT ?
                        """,
                        (chat, since_date, int(limit)),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        """
                        SELECT ts, filter_id, filter_title, item_id, item_url, title, price, region
                        FROM item_hits
                        WHERE chat_id=?
                        ORDER BY ts DESC
                        LIMIT ?
                        """,
                        (chat, int(limit)),
                    ).fetchall()
            finally:
                conn.close()
            return [dict(r) for r in rows]
        except Exception as err:
            logger.debug("stats get_item_rows failed: {}", str(err)[:160])
            return []

    def export_xlsx(self, chat_id: str, out_path: str | Path, *, days: int = 30) -> Path:
        """
        Build XLSX with 2 sheets:
        - Daily: per-day counters
        - Items: recent item hits (what items we reached)
        """
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        if days and days > 0:
            since = (datetime.now(tz=timezone.utc).date() - timedelta(days=int(days) - 1)).isoformat()
        else:
            since = None

        if not self.enabled or not self.db_path:
            wb = Workbook(write_only=True)
            ws_daily = wb.create_sheet("По дням")
            ws_daily.append(
                [
                    "Дата (UTC)",
                    "Запросов всего",
                    "Успешно (OK)",
                    "Блок/капча",
                    "429 (лимит)",
                    "Ошибка прокси",
                    "Прочие ошибки",
                    "Без прокси",
                    "Свои прокси",
                    "Бесплатные прокси",
                    "Товаров (достучались)",
                    "Последняя активность (лок.)",
                ]
            )
            ws_items = wb.create_sheet("Товары")
            ws_items.append(
                [
                    "Время (лок.)",
                    "ID фильтра",
                    "Фильтр",
                    "ID объявления",
                    "Название",
                    "Цена",
                    "Ссылка",
                    "Регион",
                ]
            )
            wb.save(out)
            return out

        wb = Workbook(write_only=True)

        ws_daily = wb.create_sheet("По дням")
        ws_daily.append(
            [
                "Дата (UTC)",
                "Запросов всего",
                "Успешно (OK)",
                "Блок/капча",
                "429 (лимит)",
                "Ошибка прокси",
                "Прочие ошибки",
                "Без прокси",
                "Свои прокси",
                "Бесплатные прокси",
                "Товаров (достучались)",
                "Последняя активность (лок.)",
            ]
        )
        tz = get_localzone()
        chat = str(chat_id or "global")
        try:
            conn = self._connect()
            try:
                if since:
                    cur = conn.execute(
                        """
                        SELECT
                            date, requests_total, ok_total, blocked_total, rate_limited_total,
                            proxy_error_total, other_error_total, direct_total, user_proxy_total,
                            free_proxy_total, items_total, last_ts
                        FROM daily_stats
                        WHERE chat_id=? AND date>=?
                        ORDER BY date ASC
                        """,
                        (chat, since),
                    )
                    by_date = {str(r["date"]): r for r in cur.fetchall()}

                    start = datetime.fromisoformat(str(since)).date()
                    end = datetime.now(tz=timezone.utc).date()
                    d = start
                    while d <= end:
                        key = d.isoformat()
                        r = by_date.get(key)
                        last_local = ""
                        last_ts = int((r["last_ts"] if r else 0) or 0)
                        if last_ts:
                            try:
                                last_local = (
                                    datetime.fromtimestamp(last_ts, tz=timezone.utc)
                                    .astimezone(tz)
                                    .replace(tzinfo=None)
                                )
                            except Exception:
                                last_local = ""
                        ws_daily.append(
                            [
                                key,
                                int((r["requests_total"] if r else 0) or 0),
                                int((r["ok_total"] if r else 0) or 0),
                                int((r["blocked_total"] if r else 0) or 0),
                                int((r["rate_limited_total"] if r else 0) or 0),
                                int((r["proxy_error_total"] if r else 0) or 0),
                                int((r["other_error_total"] if r else 0) or 0),
                                int((r["direct_total"] if r else 0) or 0),
                                int((r["user_proxy_total"] if r else 0) or 0),
                                int((r["free_proxy_total"] if r else 0) or 0),
                                int((r["items_total"] if r else 0) or 0),
                                last_local,
                            ]
                        )
                        d += timedelta(days=1)
                else:
                    cur = conn.execute(
                        """
                        SELECT
                            date, requests_total, ok_total, blocked_total, rate_limited_total,
                            proxy_error_total, other_error_total, direct_total, user_proxy_total,
                            free_proxy_total, items_total, last_ts
                        FROM daily_stats
                        WHERE chat_id=?
                        ORDER BY date ASC
                        """,
                        (chat,),
                    )
                    for r in cur:
                        last_local = ""
                        last_ts = int(r["last_ts"] or 0)
                        if last_ts:
                            try:
                                last_local = (
                                    datetime.fromtimestamp(last_ts, tz=timezone.utc)
                                    .astimezone(tz)
                                    .replace(tzinfo=None)
                                )
                            except Exception:
                                last_local = ""
                        ws_daily.append(
                            [
                                r["date"],
                                int(r["requests_total"] or 0),
                                int(r["ok_total"] or 0),
                                int(r["blocked_total"] or 0),
                                int(r["rate_limited_total"] or 0),
                                int(r["proxy_error_total"] or 0),
                                int(r["other_error_total"] or 0),
                                int(r["direct_total"] or 0),
                                int(r["user_proxy_total"] or 0),
                                int(r["free_proxy_total"] or 0),
                                int(r["items_total"] or 0),
                                last_local,
                            ]
                        )
            finally:
                conn.close()
        except Exception as err:
            logger.debug("stats export daily failed: {}", str(err)[:160])

        ws_items = wb.create_sheet("Товары")
        ws_items.append(
            [
                "Время (лок.)",
                "ID фильтра",
                "Фильтр",
                "ID объявления",
                "Название",
                "Цена",
                "Ссылка",
                "Регион",
            ]
        )
        limit = 5000
        try:
            conn = self._connect()
            try:
                if since:
                    cur = conn.execute(
                        """
                        SELECT ts, filter_id, filter_title, item_id, item_url, title, price, region
                        FROM item_hits
                        WHERE chat_id=? AND date>=?
                        ORDER BY ts DESC
                        LIMIT ?
                        """,
                        (chat, since, int(limit)),
                    )
                else:
                    cur = conn.execute(
                        """
                        SELECT ts, filter_id, filter_title, item_id, item_url, title, price, region
                        FROM item_hits
                        WHERE chat_id=?
                        ORDER BY ts DESC
                        LIMIT ?
                        """,
                        (chat, int(limit)),
                    )
                for row in cur:
                    try:
                        ts_local = (
                            datetime.fromtimestamp(int(row["ts"] or 0), tz=timezone.utc)
                            .astimezone(tz)
                            .replace(tzinfo=None)
                        )
                    except Exception:
                        ts_local = ""
                    ws_items.append(
                        [
                            ts_local,
                            row["filter_id"],
                            row["filter_title"] or "",
                            row["item_id"] or "",
                            row["title"] or "",
                            row["price"] or 0,
                            row["item_url"] or "",
                            row["region"] or "",
                        ]
                    )
            finally:
                conn.close()
        except Exception as err:
            logger.debug("stats export items failed: {}", str(err)[:160])

        wb.save(out)
        return out


def export_user_stats_xlsx(
    chat_id: int | str,
    *,
    db_path: str | Path | None = None,
    out_path: str | Path | None = None,
    days: int = 30,
) -> Path:
    from paths_helper import user_stats_xlsx_path

    db = StatsDB(db_path)
    out = Path(out_path) if out_path else user_stats_xlsx_path(chat_id)
    return db.export_xlsx(str(chat_id), out, days=days)
