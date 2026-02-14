import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from stats_service import StatsDB


class TestStatsService(unittest.TestCase):
    def test_daily_counters_and_export(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "stats.db"
            db = StatsDB(db_path)

            ts = int(datetime(2026, 2, 14, 12, 0, 0, tzinfo=timezone.utc).timestamp())
            chat_id = "123"

            db.record_request(
                chat_id=chat_id,
                filter_id=1,
                filter_title="q",
                url="https://example.test/",
                proxy_kind="direct",
                status_code=200,
                outcome="ok",
                ts=ts,
            )
            db.record_request(
                chat_id=chat_id,
                filter_id=1,
                filter_title="q",
                url="https://example.test/",
                proxy_kind="user",
                status_code=403,
                outcome="blocked",
                ts=ts,
            )
            db.record_request(
                chat_id=chat_id,
                filter_id=1,
                filter_title="q",
                url="https://example.test/",
                proxy_kind="free",
                status_code=429,
                outcome="rate_limited",
                ts=ts,
            )
            db.record_request(
                chat_id=chat_id,
                filter_id=1,
                filter_title="q",
                url="https://example.test/",
                proxy_kind="user",
                status_code=None,
                outcome="proxy_error",
                error="Proxy CONNECT aborted",
                ts=ts,
            )

            db.record_item(
                chat_id=chat_id,
                filter_id=1,
                filter_title="q",
                item_id="999",
                item_url="https://www.avito.ru/999",
                title="item",
                price=100,
                region="all",
                ts=ts,
            )

            daily = db.get_daily_rows(chat_id)
            self.assertEqual(len(daily), 1)
            row = daily[0]
            self.assertEqual(row.requests_total, 4)
            self.assertEqual(row.ok_total, 1)
            self.assertEqual(row.blocked_total, 1)
            self.assertEqual(row.rate_limited_total, 1)
            self.assertEqual(row.proxy_error_total, 1)
            self.assertEqual(row.other_error_total, 0)
            self.assertEqual(row.items_total, 1)
            self.assertEqual(row.direct_total, 1)
            self.assertEqual(row.user_proxy_total, 2)
            self.assertEqual(row.free_proxy_total, 1)

            out = Path(td) / "stats.xlsx"
            db.export_xlsx(chat_id, out, days=30)
            self.assertTrue(out.exists())

            wb = load_workbook(out)
            self.assertIn("По дням", wb.sheetnames)
            self.assertIn("Товары", wb.sheetnames)

            ws_daily = wb["По дням"]
            self.assertGreaterEqual(ws_daily.max_row, 2)
            headers = [c.value for c in next(ws_daily.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(
                headers,
                [
                    "Дата (UTC)",
                    "Запросов всего",
                    "Успешно (OK)",
                    "Блок/капча",
                    "429 (лимит)",
                    "Ошибка прокси",
                    "Прочие ошибки",
                    "Без прокси",
                    "Свои прокси",
                    "Бесплатные прокси",
                    "Товаров (достучались)",
                    "Последняя активность (лок.)",
                ],
            )

            ws_items = wb["Товары"]
            self.assertGreaterEqual(ws_items.max_row, 2)
