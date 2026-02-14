import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from threading import Lock
from datetime import datetime
from tzlocal import get_localzone
from models import Item


class XLSXHandler:
    """Сохраняет информацию в xlsx"""

    def __init__(self, file_name):
        self._initialize(file_name=file_name)

    def _initialize(self, file_name):
        self.file_name = file_name
        Path(self.file_name).parent.mkdir(parents=True, exist_ok=True)
        if not os.path.exists(self.file_name):
            self._create_file()

    def _create_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append([
            "Фильтр",
            "Регион",
            "Название",
            "Цена",
            "URL",
            "Описание",
            "Дата публикации",
            "Продавец",
            "Адрес",
            "Адрес пользователя",
            "Координаты",
            "Изображения",
            "Поднято",
            "Просмотры (всего)",
            "Просмотры (сегодня)",
            "Интервал (сек)",
        ])
        workbook.save(self.file_name)

    @staticmethod
    def get_ad_time(ad: Item):
        stamp = getattr(ad, "sortTimeStamp", None)
        if not stamp:
            return datetime.now(tz=get_localzone()).replace(tzinfo=None)
        return datetime.fromtimestamp(stamp / 1000, tz=get_localzone()).replace(tzinfo=None)

    @staticmethod
    def get_item_coords(ad: Item) -> str:
        """
        Возвращает строку с координатами "lat;lng" из Item.
        Если координаты отсутствуют — возвращает пустую строку.
        """
        if ad.coords and 'lat' in ad.coords and 'lng' in ad.coords:
            return f"{ad.coords['lat']};{ad.coords['lng']}"
        return ""

    @staticmethod
    def get_item_address_user(ad: Item) -> str:
        """
        Возвращает строку address_user из Item.
        Если address_user отсутствует — возвращает пустую строку.
        """
        if ad.coords and 'address_user' in ad.coords:
            return ad.coords['address_user']
        return ""

    def append_data_from_page(self, ads: list[Item]):
        workbook = load_workbook(self.file_name)
        sheet = workbook.active

        def get_largest_image_url(img):
            if not hasattr(img, "root"):
                return ""
            best_key = max(
                img.root.keys(),
                key=lambda k: int(k.split("x")[0]) * int(k.split("x")[1])
            )
            return str(img.root[best_key])

        for ad in ads:
            images = getattr(ad, "images", []) or []
            images_urls = [get_largest_image_url(img) for img in images if getattr(img, "root", None)]
            filter_name = getattr(ad, "filter_title", "") or ""
            region_label = getattr(ad, "filter_region_label", "") or (ad.location.name if ad.location else "")
            interval = getattr(ad, "filter_interval_seconds", None)

            row = [
                filter_name,
                region_label,
                ad.title,
                ad.priceDetailed.value,
                f"https://www.avito.ru/{ad.urlPath}",
                ad.description,
                self.get_ad_time(ad=ad),
                ad.sellerId if ad.sellerId else "",
                ad.location.name if ad.location else "",
                self.get_item_address_user(ad=ad),
                self.get_item_coords(ad=ad),
                ";".join(images_urls),
                "Да" if ad.isPromotion else "Нет",
                ad.total_views if ad.total_views is not None else "",
                ad.today_views if ad.today_views is not None else "",
                interval if interval is not None else "",
            ]
            sheet.append(row)

        workbook.save(self.file_name)
